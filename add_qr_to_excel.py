from __future__ import annotations

import argparse
from io import BytesIO
from pathlib import Path
from typing import Dict, List

import qrcode
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


def extract_url(cell) -> str | None:
    """Return a URL from cell hyperlink target or text value."""
    if cell.hyperlink and cell.hyperlink.target:
        return str(cell.hyperlink.target)

    value = cell.value
    if isinstance(value, str) and value.startswith(("http://", "https://")):
        return value.strip()

    return None


def find_link_columns(ws) -> List[int]:
    cols: set[int] = set()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if extract_url(cell):
                cols.add(cell.column)
    return sorted(cols)


def make_qr_image(url: str, size: int = 84) -> XLImage:
    qr = qrcode.QRCode(border=1, box_size=10)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    png = BytesIO()
    img.save(png, format="PNG")
    png.seek(0)

    xl_img = XLImage(png)
    xl_img.width = size
    xl_img.height = size
    return xl_img


def add_qr_columns(input_path: Path, output_path: Path) -> Dict[str, int]:
    wb = load_workbook(input_path)
    summary: Dict[str, int] = {}

    for ws in wb.worksheets:
        link_cols = find_link_columns(ws)
        if not link_cols:
            continue

        qr_col_map: Dict[int, int] = {}
        next_col = ws.max_column + 1
        for link_col in link_cols:
            qr_col_map[link_col] = next_col
            source_header = ws.cell(row=1, column=link_col).value or f"Col {link_col}"
            ws.cell(row=1, column=next_col, value=f"QR - {source_header}")
            ws.column_dimensions[get_column_letter(next_col)].width = 16
            next_col += 1

        added = 0
        for row in range(2, ws.max_row + 1):
            row_had_qr = False
            for link_col, qr_col in qr_col_map.items():
                cell = ws.cell(row=row, column=link_col)
                url = extract_url(cell)
                if not url:
                    continue

                qr_img = make_qr_image(url)
                ws.add_image(qr_img, f"{get_column_letter(qr_col)}{row}")
                added += 1
                row_had_qr = True

            if row_had_qr:
                ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, 68)

        summary[ws.title] = added

    wb.save(output_path)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Excel dosyasındaki web linkleri için otomatik QR kod sütunları ekler."
    )
    parser.add_argument("input", type=Path, help="Girdi .xlsx dosya yolu")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Çıktı .xlsx dosya yolu (verilmezse *_qr.xlsx)",
    )

    args = parser.parse_args()
    if not args.input.exists():
        raise FileNotFoundError(f"Dosya bulunamadı: {args.input}")

    if args.output:
        out = args.output
    else:
        out = args.input.with_name(f"{args.input.stem}_qr{args.input.suffix}")

    summary = add_qr_columns(args.input, out)
    total = sum(summary.values())

    print(f"Tamamlandı: {out}")
    print(f"Toplam QR: {total}")
    for sheet, count in summary.items():
        print(f"- {sheet}: {count}")


if __name__ == "__main__":
    main()
